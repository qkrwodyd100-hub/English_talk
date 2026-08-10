from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXPECTED_HEADERS = ('주차', '일차', '주제', '영어 표현', '한글 해석')
DAY_PATTERN = re.compile(r'^Day ([1-9]|[1-5][0-9]|60)$')
NUMBERING_PATTERN = re.compile(r'^\s*\d+\.\s*')


def text(value: Any) -> str:
    return '' if value is None else str(value).strip()


def extract_sentences(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    if workbook.sheetnames != ['Sheet1']:
        raise ValueError(f'expected exactly one Sheet1 worksheet, found {workbook.sheetnames!r}')

    worksheet = workbook['Sheet1']
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = tuple(text(value) for value in next(rows))
    except StopIteration as error:
        raise ValueError('workbook is empty') from error
    if headers != EXPECTED_HEADERS:
        raise ValueError(f'unexpected headers: {headers!r}')

    sentences: list[dict[str, object]] = []
    seen_english: set[str] = set()
    sentence_number_by_day: Counter[int] = Counter()
    current_day: int | None = None

    for row_number, row in enumerate(rows, start=2):
        week, day_label, topic, english, korean = (text(value) for value in row[:5])
        if not any((week, day_label, topic, english, korean)):
            continue
        if day_label:
            match = DAY_PATTERN.fullmatch(day_label)
            if not match:
                raise ValueError(f'row {row_number}: invalid day label {day_label!r}')
            current_day = int(match.group(1))
        if bool(english) != bool(korean):
            raise ValueError(f'row {row_number}: English and Korean values must both be present')
        if not english:
            continue
        if current_day is None:
            raise ValueError(f'row {row_number}: sentence has no day context')

        english = NUMBERING_PATTERN.sub('', english)
        if not english:
            raise ValueError(f'row {row_number}: English expression is blank after numbering removal')
        duplicate_key = ' '.join(english.casefold().split())
        if duplicate_key in seen_english:
            raise ValueError(f'row {row_number}: duplicate English expression {english!r}')
        seen_english.add(duplicate_key)

        sentence_number_by_day[current_day] += 1
        sentences.append({
            'id': f'day-{current_day:02d}-{sentence_number_by_day[current_day]:02d}',
            'english': english,
            'korean': korean,
            'day': current_day,
            'source': 'builtIn',
        })

    if not sentences:
        raise ValueError('workbook contains no sentences')
    return sentences


def validate_curriculum(sentences: list[dict[str, object]]) -> None:
    for index, sentence in enumerate(sentences, start=1):
        if not isinstance(sentence.get('id'), str) or not sentence['id']:
            raise ValueError(f'sentence {index}: invalid id')
        if not isinstance(sentence.get('english'), str) or not sentence['english']:
            raise ValueError(f'sentence {index}: invalid English expression')
        if not isinstance(sentence.get('korean'), str) or not sentence['korean']:
            raise ValueError(f'sentence {index}: invalid Korean translation')
        if not isinstance(sentence.get('day'), int) or not 1 <= sentence['day'] <= 60:
            raise ValueError(f'sentence {index}: invalid day')
        if sentence.get('source') != 'builtIn':
            raise ValueError(f'sentence {index}: invalid source')
    day_counts = Counter(sentence['day'] for sentence in sentences)
    expected_counts = {day: 10 for day in range(1, 61)}
    if dict(sorted(day_counts.items())) != expected_counts:
        raise ValueError(f'expected 10 sentences for each day 1-60, found {dict(sorted(day_counts.items()))!r}')


def write_typescript(sentences: list[dict[str, object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    data = json.dumps(sentences, ensure_ascii=False, indent=2)
    output_path.write_text(
        "import type { Sentence } from './learning'\n\n"
        f'export const builtInSentences: Sentence[] = {data}\n',
        encoding='utf-8',
    )


def main() -> None:
    parser = argparse.ArgumentParser(description='Convert English Talk XLSX content to static TypeScript sentences.')
    parser.add_argument('workbook', type=Path)
    parser.add_argument('--output', type=Path, default=Path('src/sentences.ts'))
    arguments = parser.parse_args()

    sentences = extract_sentences(arguments.workbook)
    validate_curriculum(sentences)
    write_typescript(sentences, arguments.output)
    print(json.dumps({
        'input': str(arguments.workbook),
        'output': str(arguments.output),
        'sentence_count': len(sentences),
        'day_counts': dict(sorted(Counter(sentence['day'] for sentence in sentences).items())),
        'samples': [sentences[0], sentences[len(sentences) // 2], sentences[-1]],
    }, ensure_ascii=False))


if __name__ == '__main__':
    main()
